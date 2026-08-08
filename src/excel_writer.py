from pathlib import Path
from typing import List, Union, Dict, Any
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from loguru import logger

from config.settings import OUTPUT_DIR
from src.schemas import ReceiptData


# ── Persian column name mapping ───────────────────────────────────────────────
# Order is optimised: transaction overview → source info → destination info → meta
COLUMN_ORDER = [
    # Transaction overview
    "amount",
    "title",
    "transaction_status",
    "transaction_datetime",
    "transfer_type",
    "tracking_id",
    "transfer_tracking_number",
    # Source side
    "source_bank",
    "source_deposit_number",
    "source_account_number",
    "source_card",
    "payer_name",
    # Destination side
    "destination_bank",
    "destination_iban",
    "destination_account_number",
    "destination_card",
    "receiver_name",
]

PERSIAN_COLUMN_NAMES: Dict[str, str] = {
    "amount":                    "مبلغ",
    "title":                     "عنوان",
    "transaction_status":        "وضعیت تراکنش",
    "transaction_datetime":      "تاریخ و ساعت تراکنش",
    "transfer_type":             "نوع انتقال",
    "tracking_id":               "شناسه پیگیری",
    "transfer_tracking_number":  "شماره پیگیری انتقال وجه",
    # Source
    "source_bank":               "بانک مبدا",
    "source_deposit_number":     "شماره سپرده مبدا",
    "source_account_number":     "شماره حساب مبدا",
    "source_card":               "شماره کارت مبدا",
    "payer_name":                "نام صاحب سپرده مبدا",
    # Destination
    "destination_bank":          "بانک مقصد",
    "destination_iban":          "شبا مقصد",
    "destination_account_number": "شماره حساب مقصد",
    "destination_card":          "شماره کارت مقصد",
    "receiver_name":             "نام صاحب شبا مقصد",
}

# Placeholder for any field that could not be identified in the receipt image
MISSING_VALUE_PLACEHOLDER = "-"
MASTER_FILENAME = "receipts.xlsx"


class ExcelExporter:
    """Exports validated receipt extraction results into a single persistent Excel file.

    Each run appends new rows to the master file (receipts.xlsx) instead of
    creating a new timestamped file. The header row is written only once.
    """

    def __init__(self, output_dir: Union[str, Path] = OUTPUT_DIR):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.master_file = self.output_dir / MASTER_FILENAME

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _receipt_to_row(self, item: Union[ReceiptData, Dict[str, Any]]) -> Dict[str, Any]:
        """Convert a single ReceiptData / dict to a flat row dict with Persian keys."""
        raw: Dict[str, Any] = item.model_dump() if isinstance(item, ReceiptData) else dict(item)

        row: Dict[str, str] = {}
        for field in COLUMN_ORDER:
            value = raw.get(field)
            if value is None or str(value).strip() == "":
                display = MISSING_VALUE_PLACEHOLDER
            else:
                display = str(value)
            row[PERSIAN_COLUMN_NAMES[field]] = display

        return row

    def _write_header(self, ws) -> None:
        """Write a styled header row to the worksheet."""
        persian_columns = [PERSIAN_COLUMN_NAMES[f] for f in COLUMN_ORDER]
        for col_idx, col_name in enumerate(persian_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="2E6DA4")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _auto_fit_columns(self, ws) -> None:
        """Adjust column widths based on cell content."""
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 55)

    # ── Public API ────────────────────────────────────────────────────────────

    def export(
        self,
        receipts: List[Union[ReceiptData, Dict[str, Any]]],
        filename_prefix: str = "receipts"   # kept for API compatibility, unused
    ) -> Path:
        """Append receipts to the master Excel file (receipts.xlsx).

        If the master file does not exist yet it is created with a styled header.
        Subsequent runs append new rows below the existing data without touching
        previously saved records.

        Args:
            receipts: List of ReceiptData instances or dictionaries.
            filename_prefix: Ignored — kept for backwards compatibility.

        Returns:
            Path to the master Excel file.
        """
        rows = [self._receipt_to_row(r) for r in receipts] if receipts else []
        persian_columns = [PERSIAN_COLUMN_NAMES[f] for f in COLUMN_ORDER]

        if not rows:
            logger.warning("No receipt data provided for export. Master file not modified.")
            # Still ensure the master file exists with headers
            if not self.master_file.exists():
                wb = Workbook()
                ws = wb.active
                ws.title = "رسیدها"
                self._write_header(ws)
                wb.save(self.master_file)
            return self.master_file

        file_path = self.master_file

        try:
            if file_path.exists():
                # ── Append mode: load existing workbook and add rows ──────────
                wb = load_workbook(file_path)
                ws = wb["رسیدها"] if "رسیدها" in wb.sheetnames else wb.active
                next_row = ws.max_row + 1
                logger.info(
                    f"Appending {len(rows)} row(s) to existing master file "
                    f"(current rows: {ws.max_row - 1}): {file_path}"
                )
            else:
                # ── Create mode: new workbook with styled header ──────────────
                wb = Workbook()
                ws = wb.active
                ws.title = "رسیدها"
                self._write_header(ws)
                next_row = 2  # row 1 is the header
                logger.info(f"Creating new master Excel file: {file_path}")

            # Write data rows
            for row_dict in rows:
                for col_idx, col_name in enumerate(persian_columns, start=1):
                    ws.cell(row=next_row, column=col_idx, value=row_dict.get(col_name, MISSING_VALUE_PLACEHOLDER))
                next_row += 1

            self._auto_fit_columns(ws)

            # Try saving to master file; fall back to timestamped file if locked
            try:
                wb.save(file_path)
                save_path = file_path
            except (PermissionError, OSError) as lock_err:
                import datetime
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                fallback_path = self.output_dir / f"receipts_{ts}.xlsx"
                logger.warning(
                    f"Master file is locked (open in Excel?): {lock_err}. "
                    f"Saving to fallback file: {fallback_path}"
                )
                wb.save(fallback_path)
                save_path = fallback_path

            logger.success(
                f"Successfully saved {len(rows)} record(s) to master file: {save_path}"
            )
            return save_path

        except Exception as e:
            logger.error(f"Failed to export to Excel master file: {e}")
            raise

